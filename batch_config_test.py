"""
Batch Configuration Testing Framework

Runs backtests across multiple YAML configurations and compares results.
Enables systematic testing of parameter variations without code changes.

Part of Configuration-Based Testing Framework (Phase 2)
"""

import pandas as pd
import os
import sys
from datetime import datetime
from typing import Dict, List
from pathlib import Path

# Import configuration loader
from config_loader import load_config, ConfigValidationError

# Import batch backtesting
import batch_backtest

# Import logging
from error_handler import setup_logging, logger

setup_logging()


def run_config_comparison(
    config_files: List[str],
    ticker_file: str,
    output_dir: str = 'backtest_results/config_comparison'
) -> Dict:
    """
    Run backtests using multiple configuration files and compare results.
    
    Args:
        config_files: List of paths to YAML configuration files
        ticker_file: Path to file containing ticker symbols
        output_dir: Directory to save comparison reports
        
    Returns:
        Dict with results for each configuration
    """
    # Create output directory
    os.makedirs(output_dir, exist_ok=True)
    
    logger.info(f"Starting configuration comparison")
    logger.info(f"Configurations to test: {len(config_files)}")
    logger.info(f"Ticker file: {ticker_file}")
    logger.info(f"Output directory: {output_dir}")
    
    # Storage for all config results
    comparison_results = {
        'configs_tested': [],
        'configs_failed': [],
        'results_by_config': {},
        'ticker_file': ticker_file,
        'timestamp': datetime.now()
    }
    
    # Test each configuration
    for i, config_path in enumerate(config_files, 1):
        logger.info(f"\n{'='*80}")
        logger.info(f"Testing configuration {i}/{len(config_files)}: {config_path}")
        logger.info(f"{'='*80}")
        
        try:
            # Load and validate configuration
            config_loader = load_config(config_path)
            config_loader.print_summary()
            
            # Extract config name for reporting
            config_name = config_loader.config.get('config_name', Path(config_path).stem)
            
            # Extract parameters for RiskManager
            risk_params = config_loader.get_risk_manager_params()
            stop_params = config_loader.get_stop_params()
            backtest_config = config_loader.get_backtest_config()
            
            # Create config-specific output directory
            config_output_dir = os.path.join(output_dir, config_name)
            os.makedirs(config_output_dir, exist_ok=True)
            
            # Run batch backtest with this configuration
            logger.info(f"Running batch backtest with {config_name} configuration...")
            
            # Convert lookback_months integer to period string (e.g., 36 -> "36mo")
            lookback_months = backtest_config.get('lookback_months', 36)
            period = f"{lookback_months}mo"
            
            results = batch_backtest.run_batch_backtest(
                ticker_file=ticker_file,
                period=period,  # Now properly formatted as string
                start_date=backtest_config.get('start_date'),
                end_date=backtest_config.get('end_date'),
                output_dir=config_output_dir,
                risk_managed=True,  # Always use risk management
                account_value=risk_params['account_value'],
                risk_pct=risk_params['risk_pct_per_trade'],
                stop_strategy=risk_params['stop_strategy'],
                time_stop_bars=risk_params['time_stop_bars'],
                save_individual_reports=False  # Save space - we'll generate aggregate comparison
            )
            
            # Store results
            comparison_results['configs_tested'].append(config_name)
            comparison_results['results_by_config'][config_name] = {
                'config_path': config_path,
                'config': config_loader.config,
                'backtest_results': results,
                'output_dir': config_output_dir
            }
            
            logger.info(f"✅ Completed {config_name}: {len(results.get('all_paired_trades', []))} trades")
            
        except ConfigValidationError as e:
            logger.error(f"❌ Configuration validation failed for {config_path}: {e}")
            comparison_results['configs_failed'].append({
                'config_path': config_path,
                'error': str(e),
                'error_type': 'validation'
            })
            
        except Exception as e:
            logger.error(f"❌ Backtest failed for {config_path}: {e}")
            comparison_results['configs_failed'].append({
                'config_path': config_path,
                'error': str(e),
                'error_type': 'execution'
            })
    
    return comparison_results


def generate_comparison_report(comparison_results: Dict, output_dir: str) -> str:
    """
    Generate comparison report across all configurations.
    
    Args:
        comparison_results: Results from run_config_comparison
        output_dir: Directory to save reports
        
    Returns:
        Path to generated report file
    """
    logger.info("\n" + "="*80)
    logger.info("Generating configuration comparison report...")
    logger.info("="*80)
    
    # Extract data for comparison table
    comparison_data = []
    
    for config_name in comparison_results['configs_tested']:
        config_result = comparison_results['results_by_config'][config_name]
        backtest_results = config_result['backtest_results']
        config = config_result['config']
        
        # Extract key metrics from risk analysis
        risk_analysis = backtest_results.get('risk_analysis', {})
        trades = backtest_results.get('all_paired_trades', [])
        
        if not trades:
            logger.warning(f"No trades for {config_name}, skipping...")
            continue
        
        # Calculate key performance metrics
        trades_df = pd.DataFrame(trades)
        
        # Basic metrics
        total_trades = len(trades)
        closed_trades = len([t for t in trades if not t.get('partial_exit', False)])
        
        # R-multiple metrics
        r_multiples = trades_df['r_multiple'].dropna()
        avg_r = r_multiples.mean() if len(r_multiples) > 0 else 0
        median_r = r_multiples.median() if len(r_multiples) > 0 else 0
        
        # Win rate
        winners = (r_multiples > 0).sum()
        win_rate = (winners / len(r_multiples) * 100) if len(r_multiples) > 0 else 0
        
        # P&L metrics
        dollar_pnl = trades_df.get('dollar_pnl', pd.Series([0] * len(trades_df)))
        total_pnl = dollar_pnl.sum()
        starting_equity = config['risk_management']['account_value']
        return_pct = (total_pnl / starting_equity * 100) if starting_equity else 0
        
        # Profit factor
        winners_pnl = dollar_pnl[dollar_pnl > 0].sum()
        losers_pnl = abs(dollar_pnl[dollar_pnl < 0].sum())
        profit_factor = (winners_pnl / losers_pnl) if losers_pnl > 0 else 0
        
        # Holding period
        bars_held = trades_df.get('bars_held', pd.Series([0] * len(trades_df)))
        avg_holding = bars_held.mean()
        
        # Drawdown (simple calculation from equity curve)
        if 'equity_after_trade' in trades_df.columns:
            equity_curve = trades_df['equity_after_trade'].dropna()
            if len(equity_curve) > 0:
                cummax = equity_curve.expanding().max()
                drawdowns = (equity_curve - cummax) / cummax * 100
                max_drawdown = drawdowns.min()
            else:
                max_drawdown = 0
        else:
            max_drawdown = 0
        
        # Exit type distribution
        exit_counts = trades_df.get('exit_type', pd.Series(['UNKNOWN'] * len(trades_df))).value_counts().to_dict()
        
        # Configuration parameters
        risk_pct = config['risk_management']['risk_pct_per_trade']
        stop_strategy = config['risk_management']['stop_strategy']
        time_stop_bars = config['risk_management']['time_stop_bars']
        
        comparison_data.append({
            'Configuration': config_name,
            'Description': config.get('description', ''),
            'Stop Strategy': stop_strategy,
            'Risk %': risk_pct,
            'Time Stop': time_stop_bars,
            'Total Trades': total_trades,
            'Closed Trades': closed_trades,
            'Win Rate %': win_rate,
            'Avg R-Multiple': avg_r,
            'Median R-Multiple': median_r,
            'Total P&L $': total_pnl,
            'Return %': return_pct,
            'Profit Factor': profit_factor,
            'Avg Holding (bars)': avg_holding,
            'Max Drawdown %': max_drawdown,
            'Time Stops': exit_counts.get('TIME_STOP', 0),
            'Hard Stops': exit_counts.get('HARD_STOP', 0),
            'Profit Targets': exit_counts.get('PROFIT_TARGET', 0),
            'Trail Stops': exit_counts.get('TRAIL_STOP', 0),
            'Signal Exits': exit_counts.get('SIGNAL_EXIT', 0)
        })
    
    # Create DataFrame for easy comparison
    comparison_df = pd.DataFrame(comparison_data)
    
    # Sort by return %
    comparison_df = comparison_df.sort_values('Return %', ascending=False)
    
    # Save to CSV
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    csv_filename = f"config_comparison_{timestamp}.csv"
    csv_path = os.path.join(output_dir, csv_filename)
    comparison_df.to_csv(csv_path, index=False)
    logger.info(f"✅ Comparison CSV saved: {csv_path}")
    
    # Save to Excel with formatting
    try:
        excel_filename = f"config_comparison_{timestamp}.xlsx"
        excel_path = os.path.join(output_dir, excel_filename)
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            comparison_df.to_excel(writer, sheet_name='Comparison', index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Comparison']
            
            # Format headers
            from openpyxl.styles import Font, Alignment, PatternFill
            
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Format numeric columns
            from openpyxl.styles import numbers
            
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                # Win Rate %
                row[7].number_format = '0.0'
                # Avg R-Multiple
                row[8].number_format = '0.00'
                # Median R-Multiple
                row[9].number_format = '0.00'
                # Total P&L $
                row[10].number_format = '$#,##0'
                # Return %
                row[11].number_format = '0.00'
                # Profit Factor
                row[12].number_format = '0.00'
                # Avg Holding
                row[13].number_format = '0.0'
                # Max Drawdown %
                row[14].number_format = '0.00'
            
            # Auto-size columns
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add conditional formatting for Return %
            from openpyxl.formatting.rule import ColorScaleRule
            
            return_col = 'L'  # Return % column
            worksheet.conditional_formatting.add(
                f'{return_col}2:{return_col}{worksheet.max_row}',
                ColorScaleRule(
                    start_type='min', start_color='F8696B',  # Red for low
                    mid_type='percentile', mid_value=50, mid_color='FFEB84',  # Yellow for mid
                    end_type='max', end_color='63BE7B'  # Green for high
                )
            )
        
        logger.info(f"✅ Comparison Excel saved: {excel_path}")
        
    except Exception as e:
        logger.warning(f"Failed to create Excel file: {e}")
        excel_path = None
    
    # Generate text report
    report_lines = []
    report_lines.append("="*100)
    report_lines.append("📊 CONFIGURATION COMPARISON REPORT")
    report_lines.append("="*100)
    report_lines.append("")
    report_lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    report_lines.append(f"Ticker File: {comparison_results['ticker_file']}")
    report_lines.append(f"Configurations Tested: {len(comparison_results['configs_tested'])}")
    report_lines.append("")
    
    if comparison_results['configs_failed']:
        report_lines.append("⚠️  FAILED CONFIGURATIONS:")
        for failure in comparison_results['configs_failed']:
            report_lines.append(f"  • {failure['config_path']}: {failure['error']}")
        report_lines.append("")
    
    report_lines.append("="*100)
    report_lines.append("🏆 RESULTS RANKING (by Return %)")
    report_lines.append("="*100)
    report_lines.append("")
    
    for rank, row in comparison_df.iterrows():
        report_lines.append(f"{rank + 1}. {row['Configuration']}")
        report_lines.append(f"   {'─'*96}")
        report_lines.append(f"   Description: {row['Description']}")
        report_lines.append(f"   Stop Strategy: {row['Stop Strategy']} | Risk: {row['Risk %']}% | Time Stop: {row['Time Stop']} bars")
        report_lines.append("")
        report_lines.append(f"   💰 Return: {row['Return %']:+.2f}% | Total P&L: ${row['Total P&L $']:,.0f}")
        report_lines.append(f"   📊 Trades: {row['Total Trades']} total ({row['Closed Trades']} closed)")
        report_lines.append(f"   ✅ Win Rate: {row['Win Rate %']:.1f}%")
        report_lines.append(f"   📈 R-Multiple: {row['Avg R-Multiple']:.2f} avg | {row['Median R-Multiple']:.2f} median")
        report_lines.append(f"   💵 Profit Factor: {row['Profit Factor']:.2f}")
        report_lines.append(f"   ⏱️  Avg Holding: {row['Avg Holding (bars)']:.1f} bars")
        report_lines.append(f"   📉 Max Drawdown: {row['Max Drawdown %']:.2f}%")
        report_lines.append("")
        report_lines.append(f"   Exit Breakdown:")
        report_lines.append(f"     • Time Stops: {row['Time Stops']}")
        report_lines.append(f"     • Hard Stops: {row['Hard Stops']}")
        report_lines.append(f"     • Profit Targets: {row['Profit Targets']}")
        report_lines.append(f"     • Trail Stops: {row['Trail Stops']}")
        report_lines.append(f"     • Signal Exits: {row['Signal Exits']}")
        report_lines.append("")
    
    # Key insights
    report_lines.append("="*100)
    report_lines.append("💡 KEY INSIGHTS")
    report_lines.append("="*100)
    report_lines.append("")
    
    # Best configuration
    best_config = comparison_df.iloc[0]
    report_lines.append(f"🏆 Best Configuration: {best_config['Configuration']}")
    report_lines.append(f"   Return: {best_config['Return %']:+.2f}% | Win Rate: {best_config['Win Rate %']:.1f}%")
    report_lines.append("")
    
    # Best win rate
    best_wr = comparison_df.loc[comparison_df['Win Rate %'].idxmax()]
    report_lines.append(f"🎯 Highest Win Rate: {best_wr['Configuration']}")
    report_lines.append(f"   Win Rate: {best_wr['Win Rate %']:.1f}% | Return: {best_wr['Return %']:+.2f}%")
    report_lines.append("")
    
    # Best R-multiple
    best_r = comparison_df.loc[comparison_df['Avg R-Multiple'].idxmax()]
    report_lines.append(f"📊 Best Avg R-Multiple: {best_r['Configuration']}")
    report_lines.append(f"   Avg R: {best_r['Avg R-Multiple']:.2f} | Return: {best_r['Return %']:+.2f}%")
    report_lines.append("")
    
    # Lowest drawdown
    lowest_dd = comparison_df.loc[comparison_df['Max Drawdown %'].idxmax()]  # Closest to 0
    report_lines.append(f"🛡️  Lowest Drawdown: {lowest_dd['Configuration']}")
    report_lines.append(f"   Max DD: {lowest_dd['Max Drawdown %']:.2f}% | Return: {lowest_dd['Return %']:+.2f}%")
    report_lines.append("")
    
    report_lines.append("="*100)
    report_lines.append("📈 PERFORMANCE COMPARISON TABLE")
    report_lines.append("="*100)
    report_lines.append("")
    report_lines.append(comparison_df.to_string(index=False))
    report_lines.append("")
    
    report_lines.append("="*100)
    report_lines.append("📝 RECOMMENDATIONS")
    report_lines.append("="*100)
    report_lines.append("")
    report_lines.append("1. Review the best performing configuration details")
    report_lines.append("2. Consider trade-offs between return and risk (drawdown)")
    report_lines.append("3. Test winning configuration on out-of-sample data")
    report_lines.append("4. Monitor consistency across different market regimes")
    report_lines.append("5. Consider combining strategies for diversification")
    report_lines.append("")
    
    report_lines.append("="*100)
    
    # Save text report
    txt_filename = f"config_comparison_{timestamp}.txt"
    txt_path = os.path.join(output_dir, txt_filename)
    
    with open(txt_path, 'w') as f:
        f.write('\n'.join(report_lines))
    
    logger.info(f"✅ Comparison report saved: {txt_path}")
    
    return txt_path


def main():
    """Main function for batch configuration testing."""
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Run batch backtesting across multiple YAML configurations'
    )
    
    parser.add_argument(
        '-c', '--configs',
        nargs='+',
        required=True,
        help='Paths to YAML configuration files (space-separated)'
    )
    
    parser.add_argument(
        '-f', '--file',
        dest='ticker_file',
        required=True,
        help='Path to file containing ticker symbols'
    )
    
    parser.add_argument(
        '-o', '--output-dir',
        default='backtest_results/config_comparison',
        help='Output directory for comparison reports (default: backtest_results/config_comparison)'
    )
    
    args = parser.parse_args()
    
    # Validate config files exist
    for config_file in args.configs:
        if not os.path.exists(config_file):
            print(f"❌ Configuration file not found: {config_file}")
            sys.exit(1)
    
    # Validate ticker file exists
    if not os.path.exists(args.ticker_file):
        print(f"❌ Ticker file not found: {args.ticker_file}")
        sys.exit(1)
    
    print(f"\n🚀 Starting batch configuration testing...")
    print(f"   Configurations: {len(args.configs)}")
    print(f"   Ticker file: {args.ticker_file}")
    print(f"   Output directory: {args.output_dir}\n")
    
    # Run comparison
    results = run_config_comparison(
        config_files=args.configs,
        ticker_file=args.ticker_file,
        output_dir=args.output_dir
    )
    
    # Generate comparison report
    report_path = generate_comparison_report(results, args.output_dir)
    
    print(f"\n{'='*80}")
    print(f"✅ Configuration comparison complete!")
    print(f"   Tested: {len(results['configs_tested'])} configurations")
    print(f"   Failed: {len(results['configs_failed'])} configurations")
    print(f"   Report: {report_path}")
    print(f"{'='*80}\n")


if __name__ == '__main__':
    main()
